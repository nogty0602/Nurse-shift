"""第1段階 看護師シフト作成 — 入力解析・スタッフマスタ・日種別."""
from openpyxl import load_workbook

# 勤務状態
OFF, LEAVE, DAY, EVE, NIGHT, OFFSITE = "OFF", "LEAVE", "DAY", "EVE", "NIGHT", "OFFSITE"
GAI, DAYNIGHT = "GAI", "DAYNIGHT"                       # 外来(0.5), 日勤深夜(ー●)
WORK_STATES = {DAY, EVE, NIGHT, OFFSITE, GAI, DAYNIGHT}     # 連勤にカウント
NIGHT_STATES = {EVE, NIGHT, DAYNIGHT}                      # 夜勤（明け・並び対象）
DEEP_STATES = {NIGHT, DAYNIGHT}                            # 深夜(3名にカウント)
DAYCOUNT_HALF = {DAY: 2, DAYNIGHT: 2, GAI: 1}             # 日勤人数の半単位換算
STATE_SYMBOL = {OFF: "×", LEAVE: "年", DAY: "ー", EVE: "▲", NIGHT: "●",
                OFFSITE: "出", GAI: "外", DAYNIGHT: "ー●"}

# セル記号 -> 固定状態
FIXED = {"×": OFF, "年": LEAVE, "ー": DAY, "▲": EVE, "●": NIGHT, "出": OFFSITE,
         "G/-": GAI, "-/G": GAI, "ケ/-": DAY, "-/ケ": DAY, "-/2": DAY}
# セル記号 -> 許容集合（除外希望）
ALLOWED = {
    "非●": {EVE, DAY, OFF}, "非▲": {NIGHT, DAY, OFF},
    "非夜": {DAY, OFF},     "非ー": {NIGHT, EVE, OFF},
    "ーor×": {DAY, OFF},    "夜or×": {NIGHT, EVE, OFF},
    "●or×": {NIGHT, OFF},   "▲or×": {EVE, OFF},
}

# スタッフマスタ: team, no_night, night_only(準夜不可), fri_mon_night(深夜は金土日月既定),
#                support_required(深夜時同チームサポート必須), can_support, no_leader, sat_night_or_off
MASTER = {
 "A": dict(team="A"), "B": dict(team="A", night_conflict=["L","S"]),
 "C": dict(team="A"), "D": dict(team="A", no_night=True),
 "E": dict(team="A"), "F": dict(team="A"),
 "G": dict(team="A", fri_mon_night=True, support_required=True),
 "H": dict(team="A"), "I": dict(team="A"), "J": dict(team="A"),
 "K": dict(team="A", night_only=True),
 "L": dict(team="A", fri_mon_night=True, support_required=True, night_conflict=["B"]),
 "M": dict(team="B"), "N": dict(team="B"), "O": dict(team="B"),
 "P": dict(team="B"), "Q": dict(team="B", night_only=True),
 "R": dict(team="B", fri_mon_night=True, support_required=True),
 "S": dict(team="B", night_conflict=["B"]),
 "T": dict(team="B", night_only=True, can_support=True, no_leader=True),
 "U": dict(team="B"),
 "V": dict(team="B", night_only=True, sat_night_or_off=True),
 "W": dict(team="B"),
 "X": dict(team="B", fri_mon_night=True, support_required=True),  # フェーズ制限は●で上書き
 "Y": dict(team="B", no_night=True),
 "Z": dict(team=None, no_night=True, chief=True),
}

DOW_FRI_MON = {"金", "土", "日", "月"}


def daytype(dow, holiday):
    if holiday:
        return "sat"
    return {"土": "sat", "日": "sun"}.get(dow, "wd")


DAY_REQ = {"wd": 10, "sat": 8, "sun": 7}
EVE_REQ = 3
NIGHT_REQ = 3


def parse(path, holidays):
    """holidays: set of day numbers (1..31) that are national holidays.
    ヘッダー行(2行目)から列位置を自動判定し、レベル・チーム・雇用・日付列を読む。"""
    wb = load_workbook(path, data_only=True)
    ws = wb["希望届"]
    # ヘッダー行を特定（「スタッフ名」を含む行）
    hdr_row = 2
    for r in range(1, 5):
        vals = [str(ws.cell(r, c).value).strip() if ws.cell(r, c).value else ""
                for c in range(1, ws.max_column + 1)]
        if "スタッフ名" in vals:
            hdr_row = r
            break
    col = {}
    day_cols = {}      # day number -> column index
    for c in range(1, ws.max_column + 1):
        v = ws.cell(hdr_row, c).value
        if v is None:
            continue
        dnum = None
        if isinstance(v, (int, float)):
            dnum = int(v)
        elif isinstance(v, str) and v.strip().isdigit():
            dnum = int(v.strip())
        if dnum is not None and 1 <= dnum <= 31:
            day_cols[dnum] = c
            continue
        key = str(v).strip()
        if "スタッフ" in key: col["name"] = c
        elif "レベル" in key or key == "Lv": col["level"] = c
        elif "チーム" in key: col["team"] = c
        elif "雇用" in key: col["emp"] = c
        elif "時短" in key: col["tanshuku"] = c
    days = sorted(day_cols)
    dow = {d: ws.cell(hdr_row + 1, day_cols[d]).value for d in days}

    staff = []
    for r in range(hdr_row + 2, ws.max_row + 1):
        name = ws.cell(r, col["name"]).value
        if name is None or str(name).strip() == "":
            continue
        lvraw = ws.cell(r, col["level"]).value if "level" in col else None
        team = ws.cell(r, col["team"]).value if "team" in col else None
        emp = ws.cell(r, col["emp"]).value if "emp" in col else None
        tan = ws.cell(r, col["tanshuku"]).value if "tanshuku" in col else None
        tanshuku = tan not in (None, "")
        cells = {}
        for d in days:
            v = ws.cell(r, day_cols[d]).value
            if v not in (None, ""):
                cells[d] = str(v).strip()
        staff.append(dict(
            name=str(name).strip(),
            level=int(lvraw) if isinstance(lvraw, (int, float)) or
                  (isinstance(lvraw, str) and lvraw.strip().isdigit()) else None,
            team=str(team).strip() if team not in (None, "") else None,
            emp=str(emp).strip() if emp not in (None, "") else None,
            tanshuku=tanshuku,
            cells=cells))
    dtype = {d: daytype(dow[d], d in holidays) for d in days}
    settings = parse_settings(wb, {s["name"] for s in staff})
    return dict(days=days, dow=dow, dtype=dtype, staff=staff, settings=settings,
                holidays=set(holidays))


def weekday_night_bounds(phase_def):
    """段階定義 -> 曜日ごとの (深夜開始前回数の下限, 上限)。None=その曜日は不可."""
    bounds = {}
    ranges = []
    lo = 1
    for p in phase_def:
        hi = p.get("cap")                    # None=以降ずっと
        ranges.append((lo, hi, p.get("weekdays", set())))
        lo = (hi + 1) if hi is not None else lo
    for w in "月火水木金土日":
        cont = [(l, h) for (l, h, wk) in ranges if w in wk]
        if not cont:
            bounds[w] = None
            continue
        pmin = min(l for l, h in cont) - 1
        his = [h for l, h in cont]
        pmax = None if any(h is None for h in his) else (max(his) - 1)
        bounds[w] = (pmin, pmax)
    return bounds


def parse_settings(wb, known_names):
    """詳細設定シートを読む: 夜勤同時不可グループ + 個人の勤務条件."""
    groups = []
    shift_rules = {}
    if "詳細設定" not in wb.sheetnames:
        return dict(night_no_overlap=groups, shift_rules=shift_rules)
    ws = wb["詳細設定"]
    rows = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            for r in range(1, ws.max_row + 1)]
    # 個人の勤務条件テーブルのヘッダー行を探す
    cond_hdr = None
    for i, row in enumerate(rows):
        cells = [str(v) for v in row if v not in (None, "")]
        if any("準夜" in c for c in cells) and any("深夜" in c for c in cells):
            cond_hdr = i
            break
    cond_rows = set()
    if cond_hdr is not None:
        for j in range(cond_hdr + 1, len(rows)):
            row = rows[j]
            nm = str(row[0]).strip() if row[0] else ""
            if nm not in known_names:
                break                                  # 空行でテーブル終了
            cond_rows.add(j)
            d = {}
            for col, key in ((1, EVE), (2, NIGHT), (3, DAY)):
                v = row[col] if len(row) > col else None
                if v not in (None, ""):
                    d[key] = str(v).strip()
            if d:
                shift_rules[nm] = d
    # 夜勤同時不可グループ（2名以上そろった行。条件表の行は除外）
    for i, row in enumerate(rows):
        if i in cond_rows:
            continue
        toks = [str(v).strip() for v in row if v not in (None, "")]
        members = [t for t in toks if t in known_names]
        hard = any(("厳守" in t or "禁止" in t or "ハード" in t) for t in toks)
        if len(members) >= 2:
            groups.append(dict(members=members, hard=hard))

    # 夜勤フェーズ定義（段階→回数上限→深夜可能曜日）
    phase_def = []
    for i, row in enumerate(rows):
        cells = [str(v).strip() for v in row if v not in (None, "")]
        if any(str(v).strip() == "段階" for v in row if v not in (None, "")) and \
           any("曜日" in str(v) for v in row if v not in (None, "")):
            for j in range(i + 1, len(rows)):
                rr = rows[j]
                stg = rr[0]
                if not (isinstance(stg, (int, float)) or (isinstance(stg, str) and str(stg).strip().isdigit())):
                    break
                cap = rr[1] if len(rr) > 1 else None
                cap = int(cap) if isinstance(cap, (int, float)) or (isinstance(cap, str) and str(cap).strip().isdigit()) else None
                wkcell = str(rr[2]).strip() if len(rr) > 2 and rr[2] else ""
                wk = [ch for ch in wkcell if ch in "月火水木金土日"]
                phase_def.append(dict(cap=cap, weekdays=set(wk)))
            break

    # レベル1 深夜経験回数（開始値・希望優先）
    lv1_exp = {}
    for i, row in enumerate(rows):
        cells = [str(v).strip() for v in row if v not in (None, "")]
        if any(str(v).strip() == "スタッフ" for v in row if v not in (None, "")) and \
           any("希望優先" in str(v) for v in row if v not in (None, "")):
            for j in range(i + 1, len(rows)):
                rr = rows[j]
                nm = str(rr[0]).strip() if rr[0] else ""
                if nm not in known_names:
                    break
                sc = rr[1] if len(rr) > 1 else 0
                sc = int(sc) if isinstance(sc, (int, float)) or (isinstance(sc, str) and str(sc).strip().isdigit()) else 0
                wp = str(rr[2]).strip() if len(rr) > 2 and rr[2] else ""
                lv1_exp[nm] = dict(start=sc, wish_priority=("○" in wp or "◯" in wp or "o" in wp.lower()))
            break

    return dict(night_no_overlap=groups, shift_rules=shift_rules,
                phase_def=phase_def, lv1_exp=lv1_exp, roster=parse_roster(rows, known_names),
                gairai=parse_gairai(rows, known_names),
                no_daynight=parse_no_daynight(rows, known_names),
                headcount=parse_headcount(rows),
                night_cap=parse_night_cap(rows, known_names))


DOW_ALL = "月火水木金土日"


def gairai_match_days(days, dow, entry):
    """外来エントリ(dow, weeks)に該当する日番号のリストを返す。weeks=None は毎週。"""
    target = entry["dow"]; weeks = entry["weeks"]
    result = []
    ordinal = 0
    for d in days:
        if dow[d] == target:
            ordinal += 1
            if weeks is None or ordinal in weeks:
                result.append(d)
    return result


def _default_req(dtype, dow_char, is_holiday):
    """必要人数(FTE)の既定値: (day_lo,day_hi, eve_lo,eve_hi, nig_lo,nig_hi)。"""
    if is_holiday or dtype == "sat":
        return (8, 8, 3, 3, 3, 3)
    if dtype == "sun":
        return (7, 7, 3, 3, 3, 3)
    if dow_char == "金":
        return (9, 9, 3, 3, 3, 3)
    if dow_char == "火":
        return (10.5, 11.5, 3, 3, 3, 3)
    return (10, 11, 3, 3, 3, 3)


def resolve_req(headcount, d, dow_char, dtype, is_holiday):
    """その日の (day,eve,nig) それぞれ [下限,上限] を返す。詳細設定＞既定。"""
    dl, dh, el, eh, nl, nh = _default_req(dtype, dow_char, is_holiday)
    req = {"day": [dl, dh], "eve": [el, eh], "nig": [nl, nh]}
    matched = []
    for r in headcount or []:
        t = str(r.get("target", "")).strip()
        pr = None
        if t == str(d):
            pr = 4
        elif t == "祝" and is_holiday:
            pr = 3
        elif t == dow_char:
            pr = 2
        elif t == "平日" and dtype == "wd" and not is_holiday:
            pr = 1
        if pr is not None:
            matched.append((pr, r))
    for pr, r in sorted(matched, key=lambda z: z[0]):
        for key, field in (("day", "day"), ("eve", "eve"), ("nig", "nig")):
            lo, hi = r.get(field, (None, None))
            if lo is not None:
                req[key][0] = lo
            if hi is not None:
                req[key][1] = hi
    return req


def parse_gairai(rows, known_names):
    """【外来割当】 曜日・時間帯(午前/午後)・対象週・担当者 を読む。"""
    out = []
    hdr = None; col = {}
    for i, row in enumerate(rows):
        cells = [str(v).strip() if v not in (None, "") else "" for v in row]
        if "曜日" in cells and "担当者" in cells:
            hdr = i
            for c, v in enumerate(cells):
                if v == "曜日": col["dow"] = c
                elif "時間帯" in v: col["ampm"] = c
                elif "対象週" in v or "週" == v: col["weeks"] = c
                elif "担当者" in v: col["staff"] = c
            break
    if hdr is None:
        return out
    for j in range(hdr + 1, len(rows)):
        row = rows[j]
        dw = str(row[col["dow"]]).strip() if len(row) > col["dow"] and row[col["dow"]] else ""
        if dw not in DOW_ALL:
            break
        ampm = str(row[col["ampm"]]).strip() if "ampm" in col and len(row) > col["ampm"] and row[col["ampm"]] else "午前"
        wtxt = str(row[col["weeks"]]).strip() if "weeks" in col and len(row) > col["weeks"] and row[col["weeks"]] else "毎週"
        staff = str(row[col["staff"]]).strip() if "staff" in col and len(row) > col["staff"] and row[col["staff"]] else ""
        weeks = None                       # None=毎週
        if "毎週" not in wtxt:
            weeks = set()
            for num, kanji in ((1, "第1"), (2, "第2"), (3, "第3"), (4, "第4"), (5, "第5")):
                if kanji in wtxt or f"第{num}" in wtxt:
                    weeks.add(num)
        pool = []
        for sep in ("・", "、", ",", "，", "/", "／", " ", "　"):
            staff = staff.replace(sep, ",")
        for t in staff.split(","):
            t = t.strip()
            if t in known_names and t not in pool:
                pool.append(t)
        out.append(dict(dow=dw, ampm=ampm, weeks=weeks, staff=pool,
                        symbol=("G/-" if "午前" in ampm else "-/G")))
    return out


def parse_no_daynight(rows, known_names):
    """【日勤深夜 不可】 スタッフ一覧 を読む。"""
    names = set()
    hdr = None
    for i, row in enumerate(rows):
        cells = [str(v).strip() if v not in (None, "") else "" for v in row]
        if any("日勤深夜" in c and "不可" in c for c in cells):
            hdr = i
            break
    if hdr is None:
        return names
    for j in range(hdr + 1, len(rows)):
        row = rows[j]
        for v in row:
            nm = str(v).strip() if v not in (None, "") else ""
            if nm in known_names:
                names.add(nm)
        # 空行で終了
        if all(v in (None, "") for v in row):
            break
    return names


def parse_headcount(rows):
    """【必要人数】 対象ごとの 日勤・準夜・深夜 下限/上限。"""
    out = []
    hdr = None; col = {}
    for i, row in enumerate(rows):
        cells = [str(v).strip() if v not in (None, "") else "" for v in row]
        if "対象" in cells and any("日勤" in c for c in cells):
            hdr = i
            for c, v in enumerate(cells):
                if v == "対象": col["target"] = c
                elif "日勤下限" in v: col["dlo"] = c
                elif "日勤上限" in v: col["dhi"] = c
                elif "準夜下限" in v: col["elo"] = c
                elif "準夜上限" in v: col["ehi"] = c
                elif "深夜下限" in v: col["nlo"] = c
                elif "深夜上限" in v: col["nhi"] = c
            break
    if hdr is None:
        return out

    def num(row, key):
        if key not in col or len(row) <= col[key]:
            return None
        try:
            return float(row[col[key]])
        except (TypeError, ValueError):
            return None

    for j in range(hdr + 1, len(rows)):
        row = rows[j]
        tgt = (str(row[col["target"]]).strip() if col.get("target") is not None
               and len(row) > col["target"] and row[col["target"]] not in (None, "") else "")
        if tgt == "":
            break
        out.append(dict(target=tgt,
                        day=(num(row, "dlo"), num(row, "dhi")),
                        eve=(num(row, "elo"), num(row, "ehi")),
                        nig=(num(row, "nlo"), num(row, "nhi"))))
    return out


def parse_night_cap(rows, known_names):
    """【夜勤上限（1人あたり月）】 対象(全員/スタッフ名) -> 上限回数。"""
    caps = {}
    hdr = None; col = {}
    for i, row in enumerate(rows):
        cells = [str(v).strip() if v not in (None, "") else "" for v in row]
        if any(c in ("対象", "スタッフ") for c in cells) and any("夜勤上限" in c for c in cells):
            hdr = i
            for c, v in enumerate(cells):
                if v in ("対象", "スタッフ"): col["name"] = c
                elif "夜勤上限" in v: col["cap"] = c
            break
    if hdr is None:
        return caps
    for j in range(hdr + 1, len(rows)):
        row = rows[j]
        nm = (str(row[col["name"]]).strip() if col.get("name") is not None
              and len(row) > col["name"] and row[col["name"]] not in (None, "") else "")
        if nm == "":
            break
        try:
            cap = int(float(row[col["cap"]]))
        except (TypeError, ValueError):
            continue
        if nm in ("全員", "すべて", "デフォルト"):
            caps["_default"] = cap
        elif nm in known_names:
            caps[nm] = cap
    return caps


def parse_roster(rows, known_names):
    """【役割設定】 スタッフ→役割 を読む。
    新形式(スタッフ|サポート|リーダー|師長) と 旧形式(スタッフ|役割) の両方に対応。"""
    roster = {}
    hdr = None
    col = {}
    for i, row in enumerate(rows):
        cells = [str(v).strip() if v not in (None, "") else "" for v in row]
        if "スタッフ" in cells and ("役割" in cells or "サポート" in cells):
            hdr = i
            for c, v in enumerate(cells):
                if v == "スタッフ": col["name"] = c
                elif v == "チーム": col["team"] = c
                elif v == "役割": col["role"] = c
                elif "サポート" in v: col["support"] = c
                elif "リーダー" in v: col["leader"] = c
                elif "師長" in v: col["chief"] = c
            break
    if hdr is None:
        return roster

    def cell(row, key):
        return (str(row[col[key]]).strip() if key in col and len(row) > col[key]
                and row[col[key]] not in (None, "") else "")

    def split_names(text):
        for sep in ("・", "、", ",", "，", "/", "／", " ", "　"):
            text = text.replace(sep, ",")
        return [t.strip() for t in text.split(",") if t.strip() in known_names]

    for j in range(hdr + 1, len(rows)):
        row = rows[j]
        members = split_names(cell(row, "name"))
        if not members:                                # 有効な名前が無ければ表の終わり
            break
        team = cell(row, "team") or None
        if "role" in col:                              # 旧形式: 役割の文字列を解釈
            roles = cell(row, "role")
            info = dict(
                team=team,
                support_required=("サポート必須" in roles),
                can_support=("サポート可" in roles or "業務可" in roles),
                no_leader=("リーダー不可" in roles),
                chief=("師長" in roles))
        else:                                          # 新形式: 列ごと
            sup = cell(row, "support"); ldr = cell(row, "leader"); chf = cell(row, "chief")
            info = dict(
                team=team,
                support_required=("必須" in sup),
                can_support=("業務可" in sup or (("可" in sup) and "必須" not in sup)),
                no_leader=("不可" in ldr),
                chief=(chf not in ("", "可能")))
        for nm in members:                             # 同じ役割を全員に適用
            roster[nm] = dict(info)
    return roster


if __name__ == "__main__":
    d = parse("/home/claude/希望届_2026年08月.xlsx", holidays={11})
    print("days:", d["days"][0], "..", d["days"][-1])
    print("daytype 8-12:", {k: d["dtype"][k] for k in range(8, 13)})
    for s in d["staff"]:
        m = MASTER.get(s["name"], {})
        print(f"{s['name']:2} Lv{s['level']} team={m.get('team')} "
              f"cells={len(s['cells'])} {list(s['cells'].items())[:4]}")
