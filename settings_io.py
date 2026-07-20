"""詳細設定の相互変換: 解析済みsettings <-> 編集用の表(行リスト) <-> Excelシート.

app.py が st.data_editor で編集した表を、parse_settings が読める『詳細設定』シート
の形式で書き戻すために使う。ヘッダー文字列は shift_core.parse_settings の検出条件に
一致させている。
"""
from openpyxl.styles import Font, PatternFill

# 各表の定義: key -> (タイトル, 注記, ヘッダー列, 見出し塗り色)
TABLE_DEFS = {
    "roles": ("【役割設定】 特別な役割の人だけ記入（チーム・レベル・雇用は希望届から読込）",
              "スタッフ欄は複数名を「・」区切りで可（例 G・L・R）。サポート: 必須/業務可 ・ リーダー: 可能/不可 ・ 師長: ○",
              ["スタッフ", "サポート", "リーダー", "師長"], "DDEBF7"),
    "overlap": ("【夜勤 同時不可グループ】 同じ日の夜勤に一緒に入れない人を1行に並べる",
                "メモは任意（自由記入）。モード欄に「厳守」＝絶対禁止 / 空欄＝回避（強いソフト）",
                ["メモ(任意)", "メンバー1", "メンバー2", "メンバー3", "メンバー4", "モード"], "DDEBF7"),
    "cond": ("【個人の勤務条件】 各シフトの可否・可能曜日を指定",
             "空欄=制限なし / 不可=禁止 / 「金土日月」=その曜日のみ / 「除日」=日曜以外",
             ["スタッフ", "準夜(▲)", "深夜(●)", "日勤(ー)", "備考"], "E2EFDA"),
    "phase": ("【夜勤フェーズ定義（レベル1・段階制）】 深夜経験の段階で入れる曜日が変わる",
              "回数上限=その段階の深夜「〜回目」まで（空欄=以降ずっと）",
              ["段階", "回数上限", "深夜可能曜日"], "FCE4D6"),
    "exp": ("【レベル1 深夜経験回数（今月開始時点）】 ここに載せた人だけ段階制を適用",
            "希望優先に ○ を入れると、解禁前の曜日でも本人の●希望を反映",
            ["スタッフ", "開始時の深夜回数", "希望優先"], "E2EFDA"),
    "gairai": ("【外来割当】 決まった曜日に担当者を外来へ（半日0.5カウント）",
               "時間帯: 午前=G/-(午前外来・午後日勤) / 午後=-/G(午前日勤・午後外来)。対象週: 毎週 / 第2・第4 など",
               ["曜日", "時間帯", "対象週", "担当者"], "FCE4D6"),
    "no_dn": ("【日勤深夜(ー●)不可】 日勤の直後に深夜へ入れない人（深夜の前は休みにする）",
              "ここに載せた人は原則ー●なし。どうしても困難な場合のみ月1回まで許容",
              ["スタッフ"], "F2DCDB"),
    "headcount": ("【必要人数】 日勤・準夜・深夜の下限/上限（対象ごと）",
                  "対象=平日/月〜日(曜日)/祝/日付(数字)。日付>祝>曜日>平日 で優先。下限=上限なら固定人数。日勤は0.5可(外来0.5換算)",
                  ["対象", "日勤下限", "日勤上限", "準夜下限", "準夜上限", "深夜下限", "深夜上限"], "FCE4D6"),
    "night_cap": ("【夜勤上限（1人あたり月）】 対象=全員 または スタッフ名",
                  "スタッフ名の行はその人だけ上書き。空なら既定10。",
                  ["対象", "夜勤上限(月)"], "E2EFDA"),
    "rest": ("【休日数（1人あたり月）】 対象=全員 または スタッフ名",
             "最低休日数=その月に必要な休みの日数（週休＋祝日分）。年休を含めるかを選べます。",
             ["対象", "最低休日数", "年休を含める"], "DDEBF7"),
    "pre_rest": ("【深夜の前は必ず休み】 深夜(●)の前日を必ず休みにする人",
                 "ここに載せた人は、深夜の前日が必ず休みになります（＝日勤深夜ー●も不可）",
                 ["スタッフ"], "F2DCDB"),
}
TABLE_ORDER = ["roles", "overlap", "cond", "phase", "exp", "gairai", "no_dn",
               "headcount", "night_cap", "rest", "pre_rest"]


def _weeks_to_text(weeks):
    if weeks is None:
        return "毎週"
    return "・".join(f"第{w}" for w in sorted(weeks))


def settings_to_rows(settings):
    """解析済み settings を、各表の行リスト(dict)に変換。編集UIの初期値に使う。"""
    out = {k: [] for k in TABLE_ORDER}

    role_groups = {}                       # (support, leader, chief) -> [names]
    for n, info in settings.get("roster", {}).items():
        support = ("サポート必須" if info.get("support_required")
                   else ("サポート業務可" if info.get("can_support") else ""))
        leader = "不可" if info.get("no_leader") else "可能"
        chief = "○" if info.get("chief") else ""
        role_groups.setdefault((support, leader, chief), []).append(n)
    for (support, leader, chief), members in role_groups.items():
        out["roles"].append(["・".join(members), support, leader, chief])

    for g in settings.get("night_no_overlap", []):
        m = list(g["members"]) + ["", "", "", ""]
        out["overlap"].append(["", m[0], m[1], m[2], m[3],
                               "厳守" if g.get("hard") else ""])

    for n, r in settings.get("shift_rules", {}).items():
        out["cond"].append([n, r.get("EVE", ""), r.get("NIGHT", ""), r.get("DAY", ""), ""])

    for i, ph in enumerate(settings.get("phase_def", []), 1):
        wk = "".join(sorted(ph["weekdays"], key="月火水木金土日".index))
        out["phase"].append([i, ph.get("cap") if ph.get("cap") is not None else "", wk])

    for n, e in settings.get("lv1_exp", {}).items():
        out["exp"].append([n, e.get("start", 0), "○" if e.get("wish_priority") else ""])

    for e in settings.get("gairai", []):
        pool = e.get("staff") or []
        if not e.get("dow") or not pool:
            continue
        out["gairai"].append([e.get("dow", ""), e.get("ampm", "午前"),
                              _weeks_to_text(e.get("weeks")), "・".join(pool)])

    for n in sorted(settings.get("no_daynight", [])):
        out["no_dn"].append([n])

    def _v(x):
        return "" if x is None else (int(x) if float(x) == int(x) else x)
    hc = settings.get("headcount", [])
    if hc:
        for r in hc:
            out["headcount"].append([r.get("target", ""),
                                     _v(r["day"][0]), _v(r["day"][1]),
                                     _v(r["eve"][0]), _v(r["eve"][1]),
                                     _v(r["nig"][0]), _v(r["nig"][1])])
    else:                                   # 既定値を初期表示（曜日ごと＋祝）
        out["headcount"] = [
            ["月", 10, 11, 3, 3, 3, 3], ["火", 10.5, 11.5, 3, 3, 3, 3],
            ["水", 10, 11, 3, 3, 3, 3], ["木", 10, 11, 3, 3, 3, 3],
            ["金", 9, 9, 3, 3, 3, 3], ["土", 8, 8, 3, 3, 3, 3],
            ["日", 7, 7, 3, 3, 3, 3], ["祝", 8, 8, 3, 3, 3, 3]]

    ncap = settings.get("night_cap", {})
    if ncap:
        if "_default" in ncap:
            out["night_cap"].append(["全員", ncap["_default"]])
        for n, c in ncap.items():
            if n != "_default":
                out["night_cap"].append([n, c])
    else:
        out["night_cap"] = [["全員", 10]]

    rest = settings.get("rest_days", {})
    if rest:
        if "_default" in rest:
            r0 = rest["_default"]
            out["rest"].append(["全員", r0["min"], "○" if r0["include_leave"] else ""])
        for n, r0 in rest.items():
            if n != "_default":
                out["rest"].append([n, r0["min"], "○" if r0["include_leave"] else ""])
    else:
        out["rest"] = [["全員", 11, "○"]]

    for n in sorted(settings.get("pre_rest", [])):
        out["pre_rest"].append([n])

    return out


def save_settings_file(rows_by_table, path):
    """編集した詳細設定だけを単独のExcelとして保存（次月に読み込む用）。"""
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    write_settings_sheet(wb, rows_by_table)
    wb.save(path)
    return path


def load_settings_file(path, known_names):
    """詳細設定ファイル（または詳細設定シートを含むExcel）から設定を読む。"""
    from openpyxl import load_workbook
    from shift_core import parse_settings
    wb = load_workbook(path, data_only=True)
    return settings_to_rows(parse_settings(wb, known_names))


def write_settings_sheet(wb, rows_by_table):
    """rows_by_table(各表の行リスト)を『詳細設定』シートとして wb に書き込む(既存は置換)。"""
    if "詳細設定" in wb.sheetnames:
        del wb["詳細設定"]
    cfg = wb.create_sheet("詳細設定")
    cfg.cell(1, 1, "詳細設定").font = Font(bold=True, size=14)
    R = 3
    for key in TABLE_ORDER:
        title, note, cols, fill = TABLE_DEFS[key]
        cfg.cell(R, 1, title).font = Font(bold=True); R += 1
        cfg.cell(R, 1, note).font = Font(italic=True, size=9); R += 1
        for c, v in enumerate(cols, 1):
            cell = cfg.cell(R, c, v); cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor=fill)
        R += 1
        for row in rows_by_table.get(key, []):
            if all(str(v).strip() == "" for v in row):
                continue
            for c, v in enumerate(row, 1):
                cfg.cell(R, c, v)
            R += 1
        R += 1
    for col, w in zip("ABCDEFGH", [16, 12, 10, 10, 10, 10, 8, 8]):
        cfg.column_dimensions[col].width = w
    return wb


def save_settings_workbook(rows_by_table, path):
    """編集した詳細設定だけを独立したExcelとして保存する（次月の読み込み用）。"""
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    write_settings_sheet(wb, rows_by_table)
    wb.save(path)
    return path


def load_settings_rows(path, known_names):
    """保存した詳細設定ファイル（詳細設定シートを含む任意のExcel）から
    編集用の行データを読み込む。"""
    from openpyxl import load_workbook
    from shift_core import parse_settings
    wb = load_workbook(path, data_only=True)
    return settings_to_rows(parse_settings(wb, known_names))
