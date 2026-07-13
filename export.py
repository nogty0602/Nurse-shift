from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from solve import solve
from shift_core import (MASTER, STATE_SYMBOL, EVE, NIGHT, DAY, OFF, LEAVE, OFFSITE,
                        GAI, DAYNIGHT, TRAIN, TRAIN_HALF, TRAIN_2H)

FILL = {  # 状態 -> (背景, 文字色)
 NIGHT: ("1F4E78", "FFFFFF"), EVE: ("E8A33D", "000000"), DAY: ("E2EFDA", "000000"),
 OFF: ("D9D9D9", "808080"), LEAVE: ("FFF2CC", "000000"), OFFSITE: ("DDEBF7", "1F4E78"),
 GAI: ("D6BFA8", "000000"), DAYNIGHT: ("8DB4E2", "1F2A44"),
 GAI if False else TRAIN: ("F4E1F0", "000000"),
 TRAIN_HALF: ("F4E1F0", "000000"), TRAIN_2H: ("F4E1F0", "000000"),
}
DOW_FILL = {"土": "DCE6F1", "日": "F2DCDB"}


def export(path, holidays, out, r=None):
    if r is None:
        r = solve(path, holidays, time_limit=90)
    roster = r["data"]["settings"].get("roster", {})
    gairai_cells = r.get("gairai_cells", {})
    A = r["assign"]; days = r["data"]["days"]; dow = r["data"]["dow"]
    names = r["names"]; lvl = r["lvl"]
    staff = {s["name"]: s for s in r["data"]["staff"]}

    wb = Workbook(); ws = wb.active; ws.title = "勤務表"
    thin = Side(style="thin", color="B0B0B0")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    ncol = 3 + len(days)
    sum_n, sum_o = ncol + 1, ncol + 2   # 夜勤数, 休日数 列

    ws.cell(1, 1, "2026年8月  勤務表（第1段階・自動生成）").font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=sum_o)

    hdr = ["スタッフ", "Lv", "チーム"] + [str(d) for d in days] + ["夜勤数", "休日数"]
    for c, v in enumerate(hdr, 1):
        cell = ws.cell(2, c, v); cell.font = Font(bold=True); cell.border = bd; cell.alignment = center
    for c, d in enumerate(days, 4):
        cell = ws.cell(3, c, dow[d]); cell.border = bd; cell.alignment = center
        if dow[d] in DOW_FILL: cell.fill = PatternFill("solid", fgColor=DOW_FILL[dow[d]])
        if d in holidays:
            cell.value = dow[d] + "祝"; cell.fill = PatternFill("solid", fgColor="F2DCDB")
    for lab, c in (("スタッフ", 1), ("Lv", 2), ("チーム", 3)):
        ws.cell(3, c, "").border = bd

    row0 = 4
    for i, n in enumerate(names):
        r0 = row0 + i
        ws.cell(r0, 1, n).font = Font(bold=True); ws.cell(r0, 1).border = bd
        ws.cell(r0, 2, lvl[n]).border = bd; ws.cell(r0, 2).alignment = center
        team = (staff[n].get("team") or roster.get(n, {}).get("team")
                or MASTER.get(n, {}).get("team") or "-")
        ws.cell(r0, 3, team).border = bd; ws.cell(r0, 3).alignment = center
        for c, d in enumerate(days, 4):
            stt = A.get((n, d), OFF); sym = STATE_SYMBOL[stt]
            if stt == DAY and staff[n].get("tanshuku"):
                sym = "P"
            elif stt == GAI:
                sym = gairai_cells.get((n, d), "外")
            cell = ws.cell(r0, c, sym); cell.border = bd; cell.alignment = center
            bg, fg = FILL[stt]
            cell.fill = PatternFill("solid", fgColor=bg)
            fixed = d in staff[n]["cells"] and staff[n]["cells"][d] not in ("",)
            cell.font = Font(color=fg, bold=fixed)  # 希望反映セルは太字
        col = get_column_letter(4); last = get_column_letter(3 + len(days))
        ws.cell(r0, sum_n, f'=COUNTIF({col}{r0}:{last}{r0},"●")+COUNTIF({col}{r0}:{last}{r0},"▲")').border = bd
        ws.cell(r0, sum_o, f'=COUNTIF({col}{r0}:{last}{r0},"×")+COUNTIF({col}{r0}:{last}{r0},"年")').border = bd
        ws.cell(r0, sum_n).alignment = center; ws.cell(r0, sum_o).alignment = center

    # 下部：日別集計（COUNTIF）
    base = row0 + len(names)
    chief_row = row0 + names.index("Z") if "Z" in names else None
    # (ラベル, 加算する記号リスト)
    labels = [("日勤(実働)", ["ー", "ー●", "P", "-/2"]), ("準夜 ▲", ["▲"]),
              ("深夜 ●", ["●", "ー●"]), ("外来", ["G/-", "-/G", "外"]), ("休 ×", ["×"])]
    top = get_column_letter(4); firstr = row0; lastr = row0 + len(names) - 1
    for k, (lab, marks) in enumerate(labels):
        rr = base + k
        ws.cell(rr, 3, lab).font = Font(bold=True); ws.cell(rr, 3).border = bd
        for c, d in enumerate(days, 4):
            L = get_column_letter(c)
            f = "+".join(f'COUNTIF({L}{firstr}:{L}{lastr},"{m}")' for m in marks)
            if "ー" in marks and chief_row:            # 日勤は師長を除外
                f += f'-COUNTIF({L}{chief_row},"ー")'
            ws.cell(rr, c, "=" + f).border = bd
            ws.cell(rr, c).alignment = center

    ws.freeze_panes = "D4"
    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 4; ws.column_dimensions["C"].width = 7
    for c in range(4, 4 + len(days)): ws.column_dimensions[get_column_letter(c)].width = 3.6
    ws.column_dimensions[get_column_letter(sum_n)].width = 7
    ws.column_dimensions[get_column_letter(sum_o)].width = 7

    # 凡例・注記シート
    ws2 = wb.create_sheet("凡例・注記")
    notes = [
        ("凡例", ""),
        ("●", "深夜"), ("▲", "準夜"), ("ー", "日勤"), ("×", "休み"),
        ("年", "年次有給"), ("出", "出張（日勤者数カウント外）"),
        ("太字セル", "希望届の希望をそのまま反映した勤務"),
        ("", ""),
        ("生成条件・前提", ""),
        ("必要人数", "日勤 平日10/土祝8/日7、準夜3、深夜3"),
        ("祝日", "8/11(山の日)は土曜扱い（8名）"),
        ("リーダー", "日勤Lv4以上1名・夜勤Lv3以上1名（T除く）"),
        ("夜勤翌日", "夜勤ブロックの後は休み（最大2連続夜勤まで）"),
        ("連勤", "最大5日"),
        ("夜勤回数", "8〜10回で均等化（D・Yは夜勤不可、Oは出張多数で6回）"),
        ("希望優先", "明示希望（●等）は曜日制限を上書き。R11(火)・X18/19も反映"),
    ]
    for rr, (a, b) in enumerate(notes, 1):
        ws2.cell(rr, 1, a).font = Font(bold=(b == "" or a in ("凡例", "生成条件・前提")))
        ws2.cell(rr, 2, b)
    ws2.column_dimensions["A"].width = 12; ws2.column_dimensions["B"].width = 55

    if r["warnings"]:
        ws3 = wb.create_sheet("警告")
        ws3.cell(1, 1, "警告").font = Font(bold=True)
        for rr, w in enumerate(r["warnings"], 2): ws3.cell(rr, 1, w)

    wb.save(out)
    print("saved", out, "status", r["status"])


if __name__ == "__main__":
    export("/home/claude/希望届_2026年08月.xlsx", {11}, "/home/claude/勤務表_2026年08月.xlsx")
