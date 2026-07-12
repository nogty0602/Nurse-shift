"""看護師シフト自動作成 — Streamlit アプリ（詳細設定を画面で編集）.

使い方:
    pip install -r requirements.txt
    streamlit run app.py
"""
import os
# ortools と streamlit の protobuf(C++)競合によるセグフォルトを回避（純Python実装に固定）
os.environ.setdefault("PROTOCOL_BUFFERS_PYTHON_IMPLEMENTATION", "python")
import re
import sys
import pickle
import subprocess
import calendar
import datetime
import tempfile
import pandas as pd
import streamlit as st
from openpyxl import load_workbook

try:
    import jpholiday
    HAS_JP = True
except ImportError:
    HAS_JP = False

# 計算は run_solver.py を別プロセスで実行するため、ここで solve/ortools は import しない
from settings_io import settings_to_rows, write_settings_sheet, TABLE_DEFS, TABLE_ORDER
from shift_core import (parse_settings, STATE_SYMBOL, OFF, DAY, EVE, NIGHT,
                        DAYNIGHT, GAI)

CELL_COLOR = {
    "●": "#1F4E78", "ー●": "#8DB4E2", "▲": "#E8A33D", "ー": "#E2EFDA",
    "P": "#E2EFDA", "×": "#D9D9D9", "年": "#FFF2CC", "出": "#DDEBF7",
    "外": "#D6BFA8", "G/-": "#D6BFA8", "-/G": "#D6BFA8", "": "#FFFFFF",
}
WHITE_TEXT = {"●", "ー●"}

st.set_page_config(page_title="看護師シフト自動作成", layout="wide")
st.title("看護師シフト自動作成")

with st.sidebar:
    st.header("入力")
    up = st.file_uploader("希望届 (.xlsx)", type=["xlsx"])
    time_limit = st.slider("計算時間の上限(秒)", 15, 180, 60, step=15)
    run = st.button("シフトを生成", type="primary", use_container_width=True)

if up is None:
    st.info("左のサイドバーで希望届(.xlsx)をアップロードしてください。"
            "『希望届』シートを含むファイルなら、詳細設定は下の画面で編集できます。")
    st.stop()

tmp_in = os.path.join(tempfile.gettempdir(), "wish_upload.xlsx")
with open(tmp_in, "wb") as f:
    f.write(up.getbuffer())
wb = load_workbook(tmp_in)
staff_names = set()
if "希望届" in wb.sheetnames:
    ws = wb["希望届"]
    for r in range(4, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v not in (None, ""):
            staff_names.add(str(v).strip())
init_rows = settings_to_rows(parse_settings(wb, staff_names))


def parse_year_month(wb):
    title = str(wb["希望届"].cell(1, 1).value or "") if "希望届" in wb.sheetnames else ""
    m = re.search(r"(\d{4})\s*年\s*(\d{1,2})\s*月", title)
    if m:
        return int(m.group(1)), int(m.group(2))
    t = datetime.date.today()
    return t.year, t.month


def auto_holidays(year, month):
    if not HAS_JP:
        return []
    n = calendar.monthrange(year, month)[1]
    return [d for d in range(1, n + 1)
            if jpholiday.is_holiday(datetime.date(year, month, d))]


# --- 対象年月・祝日 ---
st.subheader("対象年月・祝日")
y0, m0 = parse_year_month(wb)
c1, c2, c3 = st.columns([1, 1, 3])
year = c1.number_input("年", min_value=2000, max_value=2100, value=y0, step=1)
month = c2.number_input("月", min_value=1, max_value=12, value=m0, step=1)
auto = c3.checkbox("祝日を自動判定 (jpholiday)", value=HAS_JP, disabled=not HAS_JP,
                   help="希望届のタイトルから年月を読み取り、日本の祝日(振替休日含む)を自動判定します。")
if auto and HAS_JP:
    ah = auto_holidays(int(year), int(month))
    names_h = [f"{d}日({jpholiday.is_holiday_name(datetime.date(int(year),int(month),d))})" for d in ah]
    st.info("自動判定した祝日: " + ("、".join(names_h) if names_h else "なし"))
    holidays = set(ah)
else:
    txt = st.text_input("祝日(日にちをカンマ区切り)", value="")
    holidays = {int(x) for x in txt.replace("，", ",").split(",") if x.strip().isdigit()}

st.subheader("詳細設定")
st.caption("各表を直接編集できます（行の追加・削除も可）。編集後に左の「シフトを生成」を押してください。")

COND_OPTS = ["", "不可", "月火水木金", "金土日月", "土日", "除土", "除日", "除土日"]
COLCONF = {
    "roles": {"サポート": st.column_config.SelectboxColumn(
                  options=["", "サポート必須", "サポート業務可"]),
              "リーダー": st.column_config.SelectboxColumn(options=["可能", "不可"]),
              "師長": st.column_config.SelectboxColumn(options=["", "○"])},
    "cond": {"準夜(▲)": st.column_config.SelectboxColumn(options=COND_OPTS,
                 help="不可＝そのシフト禁止 / 空欄＝制限なし / 曜日＝その曜日のみ・除○＝その曜日以外"),
             "深夜(●)": st.column_config.SelectboxColumn(options=COND_OPTS),
             "日勤(ー)": st.column_config.SelectboxColumn(options=COND_OPTS)},
    "gairai": {"曜日": st.column_config.SelectboxColumn(options=["月", "火", "水", "木", "金"]),
               "時間帯": st.column_config.SelectboxColumn(options=["午前", "午後"]),
               "対象週": st.column_config.TextColumn(help="毎週 / 第2・第4 など")},
    "exp": {"希望優先": st.column_config.SelectboxColumn(options=["", "○"])},
    "overlap": {"モード": st.column_config.SelectboxColumn(options=["", "厳守"])},
    "rest": {"年休を含める": st.column_config.SelectboxColumn(options=["○", ""],
                 help="○=年休(年)も休日数に数える / 空欄=×(休)のみで数える")},
}
LABELS = {"roles": "① 役割設定", "overlap": "② 夜勤 同時不可グループ",
          "cond": "③ 個人の勤務条件", "phase": "④ 夜勤フェーズ定義",
          "exp": "⑤ レベル1 深夜経験回数", "gairai": "⑥ 外来割当",
          "no_dn": "⑦ 日勤深夜(ー●)不可", "headcount": "⑧ 必要人数(下限/上限)",
          "night_cap": "⑨ 夜勤上限(1人あたり月)", "rest": "⑩ 休日数(1人あたり月)"}

edited = {}
for key in TABLE_ORDER:
    title, note, cols, _ = TABLE_DEFS[key]
    with st.expander(LABELS[key], expanded=(key in ("roles", "gairai", "no_dn"))):
        st.caption(note)
        if key == "cond":
            st.markdown(
                "**選択肢の意味** ＝ 空欄:制限なし / 不可:そのシフト禁止 / "
                "曜日を並べたもの(例 金土日月・土日):**その曜日だけ可** / "
                "『除○』(例 除土・除日):**その曜日以外は可**\n\n"
                "例) 土曜は深夜のみ → 深夜=`除日`, 日勤=`除土`, 準夜=`不可`")
        df = pd.DataFrame(init_rows.get(key, []), columns=cols)
        ed = st.data_editor(df, num_rows="dynamic", use_container_width=True,
                            key=f"ed_{key}", column_config=COLCONF.get(key, {}))
        edited[key] = ed.values.tolist()

if run:
    write_settings_sheet(wb, edited)
    wb.save(tmp_in)

    out_xlsx = os.path.join(tempfile.gettempdir(), "schedule_out.xlsx")
    result_pkl = os.path.join(tempfile.gettempdir(), "result.pkl")
    hol_str = ",".join(str(d) for d in sorted(holidays))
    worker = os.path.join(os.path.dirname(os.path.abspath(__file__)), "run_solver.py")
    env = dict(os.environ)
    env["PROTOCOL_BUFFERS_PYTHON_IMPLEMENTATION"] = "python"

    with st.spinner("シフトを計算中…（別プロセスで実行）"):
        proc = subprocess.run(
            [sys.executable, worker, tmp_in, hol_str, str(time_limit), out_xlsx, result_pkl],
            env=env, capture_output=True, text=True, timeout=int(time_limit) + 180)

    if proc.returncode != 0:
        st.error("計算プロセスが異常終了しました（メモリ不足や制約過多の可能性）。"
                 "計算時間を短くする、制約を緩めるなどをお試しください。")
        if proc.stderr:
            with st.expander("エラー詳細"):
                st.code(proc.stderr[-2000:])
        st.stop()

    with open(result_pkl, "rb") as f:
        p = pickle.load(f)
    if not p.get("has_output"):
        st.error(f"解が見つかりませんでした (status: {p['status']})。制約が厳しすぎる可能性があります。")
        st.stop()
    st.success(f"生成完了 (status: {p['status']})")

    days = p["days"]; dow = p["dow"]; names = p["names"]; lvl = p["lvl"]; staff = p["staff"]
    A = p["assign"]; GC = p["gairai"]

    def disp(n, d):
        stt = A.get(f"{n}|{d}", OFF)
        if stt == GAI:
            return GC.get(f"{n}|{d}", "外")
        if stt == DAY and staff[n].get("tanshuku"):
            return "P"
        return STATE_SYMBOL.get(stt, "")

    grid = pd.DataFrame(
        [[disp(n, d) for d in days] for n in names],
        index=[f"{n} (Lv{lvl[n]}/{staff[n].get('team') or '-'})" for n in names],
        columns=[f"{d}\n{dow[d]}" for d in days])

    def color(v):
        bg = CELL_COLOR.get(v, "#FFFFFF")
        fg = "white" if v in WHITE_TEXT else "black"
        return f"background-color: {bg}; color: {fg}; text-align:center;"

    st.subheader("勤務表")
    styler = grid.style
    styler = styler.map(color) if hasattr(styler, "map") else styler.applymap(color)
    st.dataframe(styler, use_container_width=True, height=560)
    st.markdown("**凡例** ● 深夜 / ー● 日勤深夜 / ▲ 準夜 / ー 日勤 / P 時短 / "
                "G/-・-/G 外来 / × 休 / 年 年休 / 出 出張")

    def cnt(d, states):
        return sum(1 for n in names if A.get(f"{n}|{d}") in states
                   and staff[n].get("emp") != "師長")
    summ = pd.DataFrame({
        "日勤(実働)": [cnt(d, {DAY, DAYNIGHT}) for d in days],
        "準夜": [sum(1 for n in names if A.get(f"{n}|{d}") == EVE) for d in days],
        "深夜": [sum(1 for n in names if A.get(f"{n}|{d}") in {NIGHT, DAYNIGHT}) for d in days],
        "外来": [sum(1 for n in names if A.get(f"{n}|{d}") == GAI) for d in days],
    }, index=[f"{d}({dow[d]})" for d in days]).T
    st.subheader("日別人数")
    st.dataframe(summ, use_container_width=True)

    if p["warnings"]:
        with st.expander(f"警告 ({len(p['warnings'])}件)"):
            for w in p["warnings"]:
                st.write("・" + w)

    with open(out_xlsx, "rb") as f:
        st.download_button("勤務表(Excel)をダウンロード", f.read(),
                           file_name="勤務表.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)
